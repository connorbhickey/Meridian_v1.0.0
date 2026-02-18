"""Dialog for exporting data (CSV, JSON, Excel, chart PNG)."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

from PySide6.QtCore import Signal
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QPushButton, QComboBox, QCheckBox, QFileDialog,
    QGroupBox, QLineEdit,
)

from portopt.constants import Colors, Fonts

logger = logging.getLogger(__name__)

# Format constants
FMT_CSV_WEIGHTS = "CSV — Weights"
FMT_CSV_TRADES = "CSV — Trades"
FMT_CSV_METRICS = "CSV — Metrics"
FMT_JSON_RESULTS = "JSON — Optimization Results"
FMT_JSON_SESSION = "JSON — Full Session State"
FMT_EXCEL_REPORT = "Excel — Portfolio Report"
FMT_PNG_CHARTS = "PNG — All Charts"

_FILTER_MAP = {
    FMT_CSV_WEIGHTS: "CSV Files (*.csv);;All Files (*)",
    FMT_CSV_TRADES: "CSV Files (*.csv);;All Files (*)",
    FMT_CSV_METRICS: "CSV Files (*.csv);;All Files (*)",
    FMT_JSON_RESULTS: "JSON Files (*.json);;All Files (*)",
    FMT_JSON_SESSION: "JSON Files (*.json);;All Files (*)",
    FMT_EXCEL_REPORT: "Excel Files (*.xlsx);;All Files (*)",
    FMT_PNG_CHARTS: "PNG Files (directory)",
}


class ExportDialog(QDialog):
    """Dialog for exporting optimization/backtest results."""

    export_requested = Signal(dict)  # export config

    def __init__(self, parent=None, has_weights=False, has_backtest=False):
        super().__init__(parent)
        self._has_weights = has_weights
        self._has_backtest = has_backtest
        self.setWindowTitle("Export Results")
        self.setMinimumSize(420, 320)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        header = QLabel("Export Results")
        header.setStyleSheet(f"color: {Colors.TEXT_PRIMARY}; font-size: 14px; font-weight: bold;")
        layout.addWidget(header)

        # Format
        format_group = QGroupBox("Export Format")
        form = QFormLayout()

        self._format_combo = QComboBox()
        self._format_combo.addItems([
            FMT_CSV_WEIGHTS, FMT_CSV_TRADES, FMT_CSV_METRICS,
            FMT_JSON_RESULTS, FMT_JSON_SESSION,
            FMT_EXCEL_REPORT, FMT_PNG_CHARTS,
        ])
        self._format_combo.currentTextChanged.connect(self._on_format_changed)
        form.addRow("Format:", self._format_combo)

        self._include_metadata = QCheckBox("Include metadata header")
        self._include_metadata.setChecked(True)
        form.addRow(self._include_metadata)

        format_group.setLayout(form)
        layout.addWidget(format_group)

        # Output path
        path_group = QGroupBox("Output")
        path_layout = QHBoxLayout()
        self._path_edit = QLineEdit()
        self._path_edit.setPlaceholderText("Select output file...")
        path_layout.addWidget(self._path_edit)

        browse_btn = QPushButton("Browse...")
        browse_btn.setFixedWidth(80)
        browse_btn.clicked.connect(self._browse)
        path_layout.addWidget(browse_btn)

        path_group.setLayout(path_layout)
        layout.addWidget(path_group)

        layout.addStretch()

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)

        export_btn = QPushButton("Export")
        export_btn.setFixedWidth(80)
        export_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Colors.ACCENT};
                color: {Colors.BG_PRIMARY};
                font-weight: bold;
                border: none;
                border-radius: 3px;
                padding: 6px 12px;
            }}
            QPushButton:hover {{ background: {Colors.ACCENT_HOVER}; }}
        """)
        export_btn.clicked.connect(self._submit)
        btn_row.addWidget(export_btn)

        layout.addLayout(btn_row)

    def _on_format_changed(self, text: str):
        is_csv = "CSV" in text
        self._include_metadata.setEnabled(is_csv)
        if text == FMT_PNG_CHARTS:
            self._path_edit.setPlaceholderText("Select output directory...")
        else:
            self._path_edit.setPlaceholderText("Select output file...")

    def _browse(self):
        fmt = self._format_combo.currentText()
        if fmt == FMT_PNG_CHARTS:
            path = QFileDialog.getExistingDirectory(self, "Select Chart Export Directory")
        else:
            file_filter = _FILTER_MAP.get(fmt, "All Files (*)")
            path, _ = QFileDialog.getSaveFileName(self, "Export", "", file_filter)
        if path:
            self._path_edit.setText(path)

    def _submit(self):
        path = self._path_edit.text().strip()
        if not path:
            return

        config = {
            "format": self._format_combo.currentText(),
            "path": path,
            "include_metadata": self._include_metadata.isChecked(),
        }
        self.export_requested.emit(config)
        self.accept()


# ── CSV Exporters ─────────────────────────────────────────────────────


def export_weights_csv(weights: dict[str, float], path: str,
                       metadata: dict | None = None):
    """Write weights to CSV file."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        if metadata:
            for k, v in metadata.items():
                writer.writerow([f"# {k}", v])
        writer.writerow(["Symbol", "Weight"])
        for sym, w in sorted(weights.items(), key=lambda x: -x[1]):
            writer.writerow([sym, f"{w:.6f}"])
    logger.info("Exported weights to %s", path)


def export_trades_csv(trades: list[dict], path: str):
    """Write trade blotter to CSV file."""
    if not trades:
        return
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=trades[0].keys())
        writer.writeheader()
        writer.writerows(trades)
    logger.info("Exported %d trades to %s", len(trades), path)


def export_metrics_csv(metrics: dict[str, float], path: str):
    """Write metrics to CSV file."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Metric", "Value"])
        for k, v in metrics.items():
            writer.writerow([k, f"{v:.6f}" if isinstance(v, float) else v])
    logger.info("Exported metrics to %s", path)


# ── JSON Exporters ────────────────────────────────────────────────────


def export_optimization_json(result, path: str, metadata: dict | None = None):
    """Export optimization result to JSON.

    Args:
        result: OptimizationResult with weights, metrics, etc.
        path: Output file path.
        metadata: Additional metadata to include.
    """
    data = {
        "method": str(result.method),
        "weights": {k: round(v, 6) for k, v in sorted(result.weights.items(), key=lambda x: -x[1])},
        "expected_return": round(result.expected_return, 6),
        "volatility": round(result.volatility, 6),
        "sharpe_ratio": round(result.sharpe_ratio, 6),
    }
    if result.metadata:
        # Filter to JSON-serializable values
        safe_meta = {}
        for k, v in result.metadata.items():
            try:
                json.dumps(v)
                safe_meta[k] = v
            except (TypeError, ValueError):
                safe_meta[k] = str(v)
        data["metadata"] = safe_meta

    if metadata:
        data["export_metadata"] = metadata

    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    logger.info("Exported optimization JSON to %s", path)


def export_session_json(
    opt_result=None,
    bt_output=None,
    portfolio=None,
    path: str = "",
):
    """Export full session state to JSON for reproducibility.

    Args:
        opt_result: OptimizationResult or None.
        bt_output: BacktestOutput or None.
        portfolio: Portfolio or None.
        path: Output file path.
    """
    from datetime import date, datetime

    def _safe_val(v):
        if isinstance(v, (date, datetime)):
            return v.isoformat()
        if isinstance(v, float):
            return round(v, 6)
        try:
            json.dumps(v)
            return v
        except (TypeError, ValueError):
            return str(v)

    session = {"version": "1.0"}

    if portfolio:
        session["portfolio"] = {
            "total_value": round(portfolio.total_value, 2),
            "n_holdings": len(portfolio.holdings),
            "symbols": portfolio.symbols,
            "last_updated": portfolio.last_updated.isoformat() if portfolio.last_updated else None,
        }

    if opt_result:
        session["optimization"] = {
            "method": str(opt_result.method),
            "weights": {k: round(v, 6) for k, v in opt_result.weights.items()},
            "expected_return": round(opt_result.expected_return, 6),
            "volatility": round(opt_result.volatility, 6),
            "sharpe_ratio": round(opt_result.sharpe_ratio, 6),
        }

    if bt_output and bt_output.metrics:
        session["backtest"] = {
            k: _safe_val(v) for k, v in bt_output.metrics.items()
        }

    with open(path, "w") as f:
        json.dump(session, f, indent=2)
    logger.info("Exported session state to %s", path)


# ── Excel Exporter ────────────────────────────────────────────────────


def export_excel_report(
    weights: dict[str, float] | None = None,
    metrics: dict[str, float] | None = None,
    trades: list[dict] | None = None,
    path: str = "",
):
    """Export a multi-sheet Excel workbook with portfolio data.

    Requires openpyxl. If not installed, falls back to CSV.

    Args:
        weights: Optimization weights {symbol: weight}.
        metrics: Backtest/portfolio metrics.
        trades: Trade blotter entries.
        path: Output .xlsx file path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV export")
        if weights:
            export_weights_csv(weights, path.replace(".xlsx", "_weights.csv"))
        if metrics:
            export_metrics_csv(metrics, path.replace(".xlsx", "_metrics.csv"))
        return

    wb = Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0A3D5C", end_color="0A3D5C", fill_type="solid")

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Weights sheet
    if weights:
        ws = wb.active
        ws.title = "Weights"
        ws.append(["Symbol", "Weight", "Weight %"])
        for sym, w in sorted(weights.items(), key=lambda x: -x[1]):
            ws.append([sym, round(w, 6), f"{w * 100:.2f}%"])
        _style_header(ws)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
    else:
        ws = wb.active
        ws.title = "Info"
        ws.append(["No optimization data available"])

    # Metrics sheet
    if metrics:
        ws = wb.create_sheet("Metrics")
        ws.append(["Metric", "Value"])
        for k, v in metrics.items():
            ws.append([k, round(v, 6) if isinstance(v, float) else v])
        _style_header(ws)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    # Trades sheet
    if trades:
        ws = wb.create_sheet("Trades")
        if trades:
            headers = list(trades[0].keys())
            ws.append(headers)
            for t in trades:
                ws.append([t.get(h, "") for h in headers])
            _style_header(ws)
            for col_letter in ["A", "B", "C", "D", "E"]:
                ws.column_dimensions[col_letter].width = 15

    wb.save(path)
    logger.info("Exported Excel report to %s", path)


# ── Chart PNG Exporter ────────────────────────────────────────────────


def export_charts_png(panels: dict, output_dir: str):
    """Export all pyqtgraph chart panels to PNG images.

    Args:
        panels: Dict of {name: panel} where panel has a pyqtgraph PlotWidget.
        output_dir: Directory to write PNG files into.
    """
    import pyqtgraph as pg
    from pyqtgraph.exporters import ImageExporter

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    exported = 0
    for name, panel in panels.items():
        # Find all PlotWidget children in the panel
        plot_widgets = panel.findChildren(pg.PlotWidget)
        for i, pw in enumerate(plot_widgets):
            suffix = f"_{i}" if len(plot_widgets) > 1 else ""
            filename = out / f"{name}{suffix}.png"
            try:
                exporter = ImageExporter(pw.plotItem)
                exporter.parameters()["width"] = 1200
                exporter.export(str(filename))
                exported += 1
            except Exception as e:
                logger.warning("Failed to export chart %s: %s", name, e)

    logger.info("Exported %d charts to %s", exported, output_dir)
    return exported
